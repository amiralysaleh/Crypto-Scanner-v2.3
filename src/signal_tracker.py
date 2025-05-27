# -*- coding: utf-8 -*-

import json
import os
import requests
import argparse
from datetime import datetime, timedelta
import pytz
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from filelock import FileLock
import time
import logging

from config import SIGNALS_FILE, KUCOIN_BASE_URL, KUCOIN_KLINE_ENDPOINT, KUCOIN_TICKER_ENDPOINT
from telegram_sender import send_telegram_message

# تنظیم لاگ برای دیباگ
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def load_signals():
    """Load signals from JSON file with proper timezone handling and validation."""
    lock = FileLock(f"{SIGNALS_FILE}.lock")
    try:
        with lock:
            if not os.path.exists(SIGNALS_FILE):
                logger.warning(f"No signals file found at {SIGNALS_FILE}. Creating an empty one.")
                os.makedirs(os.path.dirname(SIGNALS_FILE), exist_ok=True)
                with open(SIGNALS_FILE, 'w') as f:
                    json.dump([], f, indent=4)
                return []

            with open(SIGNALS_FILE, 'r') as f:
                content = f.read().strip()
                if not content:
                    logger.warning(f"Signals file {SIGNALS_FILE} is empty. Initializing with empty list.")
                    with open(SIGNALS_FILE, 'w') as f:
                        json.dump([], f, indent=4)
                    return []
                signals = json.loads(content)
                logger.info(f"Loaded {len(signals)} signals from {SIGNALS_FILE}")

                tehran_tz = pytz.timezone('Asia/Tehran')
                valid_signals = []
                for signal in signals:
                    try:
                        required_fields = ['symbol', 'type', 'current_price', 'target_price', 'stop_loss', 'created_at']
                        if not all(k in signal for k in required_fields):
                            logger.error(f"Skipping invalid signal (missing fields) for {signal.get('symbol', 'unknown')}: {signal}")
                            continue

                        # اعتبارسنجی مقادیر
                        if not isinstance(signal['current_price'], (int, float)) or signal['current_price'] <= 0:
                            logger.error(f"Invalid current_price for {signal['symbol']}: {signal['current_price']}")
                            continue
                        if not isinstance(signal['target_price'], (int, float)) or signal['target_price'] <= 0:
                            logger.error(f"Invalid target_price for {signal['symbol']}: {signal['target_price']}")
                            continue
                        if not isinstance(signal['stop_loss'], (int, float)) or signal['stop_loss'] <= 0:
                            logger.error(f"Invalid stop_loss for {signal['symbol']}: {signal['stop_loss']}")
                            continue

                        if 'status' not in signal or signal['status'] not in ['active', 'target_reached', 'stop_loss', 'timed_out']:
                            logger.warning(f"Fixing invalid status for {signal.get('symbol', 'unknown')}")
                            signal['status'] = 'active'

                        created_at_str = signal['created_at']
                        try:
                            if 'T' in created_at_str and ('+' in created_at_str or 'Z' in created_at_str):
                                created_at = datetime.fromisoformat(created_at_str.replace('Z', '+00:00')).astimezone(tehran_tz)
                            else:
                                created_at = tehran_tz.localize(datetime.strptime(created_at_str, "%Y-%m-%d %H:%M:%S"))
                            signal['created_at'] = created_at.isoformat()
                        except ValueError as e:
                            logger.error(f"Could not parse created_at for {signal['symbol']}: {created_at_str}. Skipping. Error: {e}")
                            continue

                        if 'closed_at' in signal and signal['closed_at']:
                            closed_at_str = signal['closed_at']
                            try:
                                if 'T' in closed_at_str and ('+' in closed_at_str or 'Z' in closed_at_str):
                                    closed_at = datetime.fromisoformat(closed_at_str.replace('Z', '+00:00')).astimezone(tehran_tz)
                                else:
                                    closed_at = tehran_tz.localize(datetime.strptime(closed_at_str, "%Y-%m-%d %H:%M:%S"))
                                signal['closed_at'] = closed_at.isoformat()
                            except ValueError:
                                signal['closed_at'] = None

                        valid_signals.append(signal)
                    except Exception as e:
                        logger.error(f"Error processing signal for {signal.get('symbol', 'unknown')}: {e}. Skipping.")
                logger.info(f"Validated {len(valid_signals)} signals")
                return valid_signals
    except json.JSONDecodeError as e:
        logger.error(f"Error decoding JSON from {SIGNALS_FILE}: {e}. File may be corrupted. Initializing with empty list.")
        with open(SIGNALS_FILE, 'w') as f:
            json.dump([], f, indent=4)
        return []
    except Exception as e:
        logger.error(f"Unexpected error loading signals: {e}")
        return []

def save_signals(signals):
    """Save signals to JSON file with proper timezone handling and file lock."""
    lock = FileLock(f"{SIGNALS_FILE}.lock")
    try:
        with lock:
            os.makedirs(os.path.dirname(SIGNALS_FILE), exist_ok=True)
            with open(SIGNALS_FILE, 'w') as f:
                json.dump(signals, f, indent=4)
            logger.info(f"Saved {len(signals)} signals to {SIGNALS_FILE}")
    except Exception as e:
        logger.error(f"Error saving signals: {e}")
        send_telegram_message(f"❌ Error saving signals to {SIGNALS_FILE}: {e}")

def save_signal(signal):
    """Save a single signal with proper timezone handling."""
    tehran_tz = pytz.timezone('Asia/Tehran')
    if 'entry_price' not in signal or not signal['entry_price']:
        signal['entry_price'] = signal.get('current_price')
    signal['status'] = 'active'
    if 'created_at' not in signal or not signal['created_at']:
        signal['created_at'] = datetime.now(tehran_tz).isoformat()
    elif isinstance(signal['created_at'], str) and 'T' not in signal['created_at']:
        try:
            dt_obj = datetime.strptime(signal['created_at'], "%Y-%m-%d %H:%M:%S")
            signal['created_at'] = tehran_tz.localize(dt_obj).isoformat()
        except ValueError:
            signal['created_at'] = datetime.now(tehran_tz).isoformat()

    signals = load_signals()
    signals.append(signal)
    save_signals(signals)
    logger.info(f"Signal saved: {signal['symbol']} {signal['type']}")

def fetch_kline_data(symbol, start_time, end_time, interval="30min"):
    """Fetch kline data from KuCoin for a specific time range with retry logic."""
    url = f"{KUCOIN_BASE_URL}{KUCOIN_KLINE_ENDPOINT}"
    params = {
        "symbol": symbol,
        "type": interval,
        "startAt": int(start_time.timestamp()),
        "endAt": int(end_time.timestamp())
    }
    for attempt in range(3):
        try:
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            if not data.get('data'):
                logger.warning(f"No kline data for {symbol} from {start_time} to {end_time}: {data}")
                return None
            df = pd.DataFrame(data['data'], columns=[
                "timestamp", "open", "close", "high", "low", "volume", "turnover"
            ])
            df = df[["timestamp", "open", "high", "low", "close", "volume"]]
            df = df.astype(float)
            df["timestamp"] = pd.to_datetime(df["timestamp"], unit="s").dt.tz_localize('UTC').dt.tz_convert('Asia/Tehran')
            df = df.iloc[::-1].reset_index(drop=True)
            logger.info(f"Received {len(df)} candles for {symbol} from {start_time} to {end_time}")
            return df
        except requests.RequestException as e:
            logger.error(f"Attempt {attempt + 1} failed for {symbol}: {e}. Retrying in {2 ** attempt}s...")
            time.sleep(2 ** attempt)
    logger.error(f"Failed to fetch kline data for {symbol} after 3 attempts")
    return None

def get_current_price(symbol):
    """Fetch current price from KuCoin with retry logic."""
    url = f"{KUCOIN_BASE_URL}{KUCOIN_TICKER_ENDPOINT}"
    params = {"symbol": symbol}
    for attempt in range(3):
        try:
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            price = data.get('data', {}).get('price')
            if price:
                return float(price)
            logger.warning(f"No price data for {symbol}: {data}")
            return None
        except requests.RequestException as e:
            logger.error(f"Attempt {attempt + 1} failed for {symbol}: {e}. Retrying in {2 ** attempt}s...")
            time.sleep(2 ** attempt)
    logger.error(f"Failed to fetch price for {symbol} after 3 attempts")
    return None

def calculate_profit_loss(signal, close_price):
    """Calculate profit/loss percentage."""
    try:
        entry_price = float(signal.get('entry_price', signal['current_price']))
        close_price = float(close_price)
        if entry_price == 0:
            logger.warning(f"Entry price is zero for {signal.get('symbol', 'N/A')}")
            return None
        if signal['type'] == 'BUY':
            return ((close_price - entry_price) / entry_price) * 100
        else:  # SELL
            return ((entry_price - close_price) / entry_price) * 100
    except (ValueError, TypeError) as e:
        logger.error(f"Error calculating profit/loss for {signal.get('symbol', 'N/A')}: {e}")
        return None

def calculate_duration(created_at_str, closed_at_str):
    """Calculate signal duration in hours."""
    tehran_tz = pytz.timezone('Asia/Tehran')
    try:
        created = datetime.fromisoformat(created_at_str).astimezone(tehran_tz)
        closed = datetime.fromisoformat(closed_at_str).astimezone(tehran_tz) if closed_at_str else datetime.now(tehran_tz)
        duration_hours = (closed - created).total_seconds() / 3600
        return duration_hours
    except (ValueError, TypeError) as e:
        logger.error(f"Error calculating duration for {created_at_str}: {e}")
        return None

def check_signal_hit(signal, df):
    """Check if signal hit target or stop-loss based on kline data."""
    try:
        target_price = float(signal['target_price'])
        stop_loss = float(signal['stop_loss'])
        signal_type = signal['type']
        created_at = datetime.fromisoformat(signal['created_at']).astimezone(pytz.timezone('Asia/Tehran'))

        for _, row in df.iterrows():
            candle_time = row['timestamp']
            if candle_time <= created_at:
                continue
            high = row['high']
            low = row['low']
            if signal_type == 'BUY':
                if high >= target_price:
                    return 'target_reached', str(target_price), candle_time.isoformat()
                if low <= stop_loss:
                    return 'stop_loss', str(stop_loss), candle_time.isoformat()
            elif signal_type == 'SELL':
                if low <= target_price:
                    return 'target_reached', str(target_price), candle_time.isoformat()
                if high >= stop_loss:
                    return 'stop_loss', str(stop_loss), candle_time.isoformat()
        return None, None, None
    except Exception as e:
        logger.error(f"Error checking signal hit for {signal['symbol']}: {e}")
        return None, None, None

def update_signal_status():
    """Update signal statuses by checking historical kline data."""
    signals = load_signals()
    if not signals:
        logger.info("No signals to update")
        return

    updated = False
    tehran_tz = pytz.timezone('Asia/Tehran')
    now = datetime.now(tehran_tz)

    for signal in signals:
        if signal['status'] != 'active':
            continue

        try:
            created_at = datetime.fromisoformat(signal['created_at']).astimezone(tehran_tz)
            df = fetch_kline_data(signal['symbol'], created_at - timedelta(minutes=30), now + timedelta(minutes=30), interval="30min")
            if df is None or df.empty:
                logger.warning(f"Skipping update for {signal['symbol']} due to missing kline data")
                continue

            status, closed_price, closed_at_iso = check_signal_hit(signal, df)
            if status:
                signal['status'] = status
                signal['closed_price'] = closed_price
                signal['closed_at'] = closed_at_iso
                updated = True
                logger.info(f"Updated {signal['symbol']}: {status} at {closed_price} on {closed_at_iso}")
                send_telegram_message(
                    f"📢 **Signal Update: {signal['symbol']}**\n"
                    f"**Status:** {status.replace('_', ' ').title()}\n"
                    f"**Closed Price:** {closed_price}\n"
                    f"**Time:** {datetime.fromisoformat(closed_at_iso).strftime('%Y-%m-%d %H:%M:%S')}"
                )
            elif (now - created_at).days > 7:
                signal['status'] = 'timed_out'
                signal['closed_price'] = str(get_current_price(signal['symbol']))
                signal['closed_at'] = now.isoformat()
                updated = True
                logger.info(f"Updated {signal['symbol']}: Timed Out at {signal['closed_price']}")

        except Exception as e:
            logger.error(f"Error updating {signal['symbol']}: {e}")

    if updated:
        save_signals(signals)
        logger.info("Signals updated successfully")
    else:
        logger.info("No signals were updated")

def send_telegram_file(file_path):
    """Send file to Telegram with retry logic."""
    bot_token = os.environ.get('TELEGRAM_BOT_TOKEN')
    chat_id = os.environ.get('TELEGRAM_CHAT_ID')

    if not bot_token or not chat_id:
        logger.error("Telegram credentials not set")
        return False

    if not os.path.exists(file_path):
        logger.error(f"File {file_path} does not exist")
        return False

    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    for attempt in range(3):
        try:
            with open(file_path, 'rb') as f:
                files = {'document': (os.path.basename(file_path), f)}
                data = {'chat_id': chat_id, 'caption': '📊 Signals Report'}
                response = requests.post(url, files=files, data=data, timeout=30)
                if response.status_code == 200:
                    logger.info(f"File {file_path} sent to Telegram")
                    return True
                else:
                    logger.error(f"Attempt {attempt + 1} failed: Error sending file ({response.status_code}): {response.text}")
                    time.sleep(2 ** attempt)
        except Exception as e:
            logger.error(f"Attempt {attempt + 1} failed: Error sending file to Telegram: {e}")
            time.sleep(2 ** attempt)
    logger.error(f"Failed to send file {file_path} to Telegram after 3 attempts")
    return False

def generate_excel_report():
    """Generate Excel report with multiple sheets."""
    update_signal_status()
    signals = load_signals()
    tehran_tz = pytz.timezone('Asia/Tehran')
    now_str = datetime.now(tehran_tz).strftime("%Y%m%d_%H%M%S")
    os.makedirs('data', exist_ok=True)
    output_file = f"data/signals_report_{now_str}.xlsx"

    all_signals_data = []
    active_signals_data = []

    logger.info(f"Processing {len(signals)} signals for report...")
    for signal in signals:
        logger.info(f"Processing signal for {signal.get('symbol', 'N/A')}")
        current_price = get_current_price(signal['symbol']) if signal['status'] == 'active' else None
        if current_price and signal['status'] == 'active':
            time.sleep(0.3)

        close_price = float(signal.get('closed_price', current_price)) if signal.get('closed_price') else current_price
        profit_loss = calculate_profit_loss(signal, close_price) if close_price is not None else None
        duration = calculate_duration(signal['created_at'], signal.get('closed_at'))

        entry_price = float(signal.get('entry_price', signal['current_price'])) if signal.get('entry_price') or signal.get('current_price') else None
        logger.info(f"Entry Price: {entry_price}, Close Price: {close_price}, Profit/Loss: {profit_loss}")

        signal_row = {
            'Symbol': signal.get('symbol', ''),
            'Type': signal.get('type', ''),
            'Entry_Price': entry_price if entry_price is not None else '',
            'Target_Price': float(signal['target_price']) if signal.get('target_price') else '',
            'Stop_Loss': float(signal['stop_loss']) if signal.get('stop_loss') else '',
            'Created_At': datetime.fromisoformat(signal['created_at']).strftime('%Y-%m-%d %H:%M') if signal.get('created_at') else '',
            'Status': signal.get('status', ''),
            'Closed_Price': float(signal['closed_price']) if signal.get('closed_price') else '',
            'Closed_At': datetime.fromisoformat(signal['closed_at']).strftime('%Y-%m-%d %H:%M') if signal.get('closed_at') else '',
            'Profit_Loss_%': round(profit_loss, 2) if profit_loss is not None else '',
            'Duration_Hours': round(duration, 2) if duration is not None else '',
            'Reasons': signal.get('reasons', '').replace('✅ ', '').replace('\n', '; ')
        }
        all_signals_data.append(signal_row)
        logger.info(f"Added to all_signals_data: {signal_row}")

        if signal['status'] == 'active' and current_price is not None and entry_price is not None:
            price_change = ((current_price - entry_price) / entry_price) * 100 if entry_price != 0 else 0
            active_signal_row = {
                'Symbol': signal.get('symbol', ''),
                'Type': signal.get('type', ''),
                'Entry_Price': entry_price,
                'Current_Price': current_price,
                'Price_Change_%': round(price_change, 2),
                'Created_At': signal_row['Created_At'],
                'Reasons': signal_row['Reasons']
            }
            active_signals_data.append(active_signal_row)
            logger.info(f"Added to active_signals_data: {active_signal_row}")

    logger.info(f"Total all_signals_data: {len(all_signals_data)}, Total active_signals_data: {len(active_signals_data)}")

    if not all_signals_data:
        logger.warning("all_signals_data is empty. No data to write to Excel.")
    if not active_signals_data:
        logger.warning("active_signals_data is empty. No active data to write to Excel.")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "All Signals"
    headers = ['Symbol', 'Type', 'Entry Price', 'Target Price', 'Stop Loss', 'Created At',
               'Status', 'Closed Price', 'Closed At', 'Profit/Loss %', 'Duration Hours', 'Reasons']
    ws1.append(headers)
    for row in all_signals_data:
        row_data = [row.get(h, '') for h in headers]
        ws1.append(row_data)
        logger.debug(f"Writing row to All Signals: {row_data}")

    ws2 = wb.create_sheet("Active Signals")
    headers_active = ['Symbol', 'Type', 'Entry Price', 'Current Price', 'Price Change %', 'Created At', 'Reasons']
    ws2.append(headers_active)
    for row in active_signals_data:
        row_data = [row.get(h, '') for h in headers_active]
        ws2.append(row_data)
        logger.debug(f"Writing row to Active Signals: {row_data}")

    ws3 = wb.create_sheet("Statistics")
    ws3.append(['Metric', 'Value'])
    closed_signals = [s for s in all_signals_data if s['Status'] in ['target_reached', 'stop_loss', 'timed_out']]
    total_signals = len(all_signals_data)
    active_count = len(active_signals_data)
    target_reached = len([s for s in closed_signals if s['Status'] == 'target_reached'])
    stop_loss_signals = len([s for s in closed_signals if s['Status'] == 'stop_loss'])
    timed_out_signals = len([s for s in closed_signals if s['Status'] == 'timed_out'])
    total_closed = target_reached + stop_loss_signals + timed_out_signals
    success_rate = (target_reached / total_closed * 100) if total_closed > 0 else 0

    profits = [s['Profit_Loss_%'] for s in closed_signals if s['Profit_Loss_%'] != '']
    avg_profit = pd.Series(profits).mean() if profits else 0

    durations = [s['Duration_Hours'] for s in closed_signals if s['Duration_Hours'] != '']
    avg_duration = pd.Series(durations).mean() if durations else 0

    stats_data = [
        {'Metric': 'Total Signals', 'Value': total_signals},
        {'Metric': 'Active Signals', 'Value': active_count},
        {'Metric': 'Target Reached', 'Value': target_reached},
        {'Metric': 'Stop Loss Hit', 'Value': stop_loss_signals},
        {'Metric': 'Timed Out', 'Value': timed_out_signals},
        {'Metric': 'Total Closed', 'Value': total_closed},
        {'Metric': 'Success Rate (%)', 'Value': round(success_rate, 2)},
        {'Metric': 'Average Profit/Loss (%)', 'Value': round(avg_profit, 2) if pd.notna(avg_profit) else 0},
        {'Metric': 'Average Duration (Hours)', 'Value': round(avg_duration, 2) if pd.notna(avg_duration) else 0}
    ]
    for stat in stats_data:
        ws3.append([stat['Metric'], stat['Value']])
        logger.debug(f"Writing stat to Statistics: {stat}")

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value or ''))
                except:
                    pass
            adjusted_width = max(12, min(max_length + 2, 50))
            ws.column_dimensions[column].width = adjusted_width
            for cell in col:
                cell.border = thin_border

        ws.freeze_panes = ws['A2']

    try:
        wb.save(output_file)
        logger.info(f"Excel report generated: {output_file}")
        with open(output_file, 'rb') as f:
            content = f.read()
            if len(content) < 1000:
                logger.warning("Generated Excel file is unusually small. Content may be incomplete.")
            else:
                logger.info("Excel file content verified as non-empty.")
    except Exception as e:
        logger.error(f"Error saving Excel report: {e}")
        send_telegram_message(f"❌ Error generating Excel report: {e}")
        return

    message = (
        f"📊 **Signals Report Generated**\n\n"
        f"🟢 Active Signals: {active_count}\n"
        f"✅ Successful Signals: {target_reached}\n"
        f"❌ Failed Signals: {stop_loss_signals}\n"
        f"⏳ Timed Out: {timed_out_signals}\n"
        f"📈 **Success Rate:** {success_rate:.2f}%\n"
        f"💰 **Avg P/L:** {avg_profit:.2f}%\n"
        f"📅 Report Time: {datetime.now(tehran_tz).strftime('%Y-%m-%d %H:%M')}"
    )
    if send_telegram_message(message):
        logger.info("Telegram message sent successfully")
    else:
        logger.error("Failed to send Telegram message")

    if send_telegram_file(output_file):
        logger.info("Excel file sent to Telegram successfully")
    else:
        logger.error("Failed to send Excel file to Telegram")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Track and report signal status')
    parser.add_argument('--report', action='store_true', help='Generate and send a status report')
    parser.add_argument('--update', action='store_true', help='Only update signal status')
    args = parser.parse_args()

    try:
        if args.report:
            generate_excel_report()
        elif args.update:
            update_signal_status()
        else:
            logger.info("Please specify an action: --report or --update")
    except Exception as e:
        logger.error(f"Error in main execution: {e}")
        send_telegram_message(f"❌ System error in reporting/updating: {e}")